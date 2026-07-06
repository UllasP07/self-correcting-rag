"""Document loaders: turn a file on disk into plain text.

Enterprise data is multi-format (the whole point of the project). Milestone 1
handles the three named in the spec — PDF, Excel, Markdown/text — with simple,
transparent extractors. Milestone 2 swaps the PDF path for Unstructured/LlamaParse
to preserve tables and layout, which pypdf mangles.
"""
from __future__ import annotations

from pathlib import Path

from pypdf import PdfReader
from openpyxl import load_workbook

SUPPORTED = {".pdf", ".md", ".txt", ".xlsx"}


def load_pdf(path: Path) -> str:
    reader = PdfReader(str(path))
    pages = [page.extract_text() or "" for page in reader.pages]
    return "\n\n".join(pages)


def load_xlsx(path: Path) -> str:
    """Flatten every sheet into 'Sheet | col=val | col=val' lines.

    Naive but readable — it turns tabular data into text the embedder can index.
    """
    wb = load_workbook(str(path), read_only=True, data_only=True)
    lines: list[str] = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(h) if h is not None else "" for h in rows[0]]
        for row in rows[1:]:
            cells = [
                f"{header[i]}={row[i]}"
                for i in range(len(row))
                if i < len(header) and row[i] is not None
            ]
            if cells:
                lines.append(f"{ws.title} | " + " | ".join(cells))
    return "\n".join(lines)


def load_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")


def load_document(path: Path) -> str:
    """Dispatch on file extension. Raises for unsupported types."""
    ext = path.suffix.lower()
    if ext == ".pdf":
        return load_pdf(path)
    if ext == ".xlsx":
        return load_xlsx(path)
    if ext in {".md", ".txt"}:
        return load_text(path)
    raise ValueError(f"Unsupported file type: {ext}")
